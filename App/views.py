import numpy as np
from django.core.mail import send_mail, BadHeaderError
from django.views.decorators.csrf import csrf_exempt
from reportlab.lib import colors
from reportlab.platypus import SimpleDocTemplate, Image, Table, TableStyle, Paragraph
from reportlab.lib.styles import ParagraphStyle
from openpyxl import load_workbook
from django.contrib.staticfiles import finders
from django.utils import timezone
import matplotlib.pyplot as plt
import random
from django.core.paginator import *
from .forms import *
import cv2
import face_recognition
import datetime
import os
import pandas as pd
import string
import csv
from django.contrib.auth import authenticate, login
from django.shortcuts import render, redirect
from openpyxl.workbook import Workbook
from .models import Salle, Filiere, Cour, Professeur, Parent, Etudiant, Salle_Cour
from django.http import HttpResponse
from django.template.loader import get_template
from django.views import View
from io import BytesIO
from reportlab.pdfgen import canvas
from reportlab.pdfgen import canvas
from reportlab.lib.pagesizes import letter
from reportlab.lib.utils import ImageReader
# matplotlib.use('Agg')


fields = [
    {
        "id": 1,
        "name": "informatique et réseaux"
    },
    {
        "id": 2,
        "name": "automatismes et informatique industrielle"
    },
    {
        "id": 3,
        "name": "génie industriel"
    },
    {
        "id": 4,
        "name": "finance et audit"
    }
]



def is_time_in_range(time, start_time, end_time):
    # Convert the time strings to datetime.time objects
    time_obj = datetime.datetime.strptime(time, '%H:%M').time()
    start_time_obj = datetime.datetime.strptime(start_time, '%H:%M').time()
    end_time_obj = datetime.datetime.strptime(end_time, '%H:%M').time()
    # Check if the time is within the range
    if start_time <= time <= end_time:
        return True
    else:
        return False








def Login_user(request):
    username = request.POST.get('username')
    password = request.POST.get('password')
    user = authenticate(request, username=username, password=password)
    if user is not None:
        login(request, user)
        return redirect('etudiant_list')
    return render(request, 'login.html')

    generateClasses()
    generateFieldsList()
    courses_list()
    teachers()
    parents()
    generate_students(10)
    program_course()
    return render(request, 'login.html')


def generate_name():
    first_names = ['John', 'Jane', 'Michael', 'Emily', 'Daniel', 'Olivia']
    last_names = ['Smith', 'Johnson', 'Williams', 'Brown', 'Jones']
    first_name = random.choice(first_names)
    last_name = random.choice(last_names)
    return first_name, last_name


def generate_email(first_name, last_name):
    email_provider = ['gmail.com', 'yahoo.com', 'outlook.com']
    email = f'{first_name.lower()}.{last_name.lower()}@{random.choice(email_provider)}'
    return email


def generate_phone_number():
    return ''.join(random.choices(string.digits, k=10))


def generate_salary():
    return random.randint(20000, 50000)


def parents():
    people = []
    for _ in range(20):
        cne = generate_cne()
        first_name, last_name = generate_name()
        gender = random.choice(['M', 'F'])
        email = generate_email(first_name, last_name)
        phone_number = generate_phone_number()
        person = {
            'CNE': cne,
            'First_Name': first_name,
            'Last_Name': last_name,
            'Gender': gender,
            'Email': email,
            'Phone_Number': phone_number,
        }
        people.append(person)
        if len(Parent.objects.all()) < 20:
            t = Parent(None, person['CNE'], person['First_Name'], person['Last_Name'], person['Gender'],
                       person['Email'], person['Phone_Number'])
            t.save()


def teachers():
    people = []
    for _ in range(20):
        cne = generate_cne()
        first_name, last_name = generate_name()
        gender = random.choice(['M', 'F'])
        email = generate_email(first_name, last_name)
        phone_number = generate_phone_number()
        salary = generate_salary()

        person = {
            'CNE': cne,
            'First_Name': first_name,
            'Last_Name': last_name,
            'Gender': gender,
            'Email': email,
            'Phone_Number': phone_number,
            'Salary': salary
        }
        people.append(person)
        if len(Professeur.objects.all()) < 20:
            t = Professeur(None, person['CNE'], person['First_Name'], person['Last_Name'], person['Gender'],
                           person['Email'], person['Phone_Number'], person['Salary'])
            t.save()


def generate_students(num_students):
    def generate_parent_id():
        return random.randint(1, 20)

    def generate_level():
        return 'year' + str(random.randint(1, 5))

    def generate_group():
        return 'g' + str(random.randint(1, 3))

    students = []
    for field in fields:
        field_id = field['id']
        for _ in range(num_students):
            cne = generate_cne()
            first_name, last_name = generate_name()
            gender = random.choice(['Male', 'Female'])
            email = generate_email(first_name, last_name)
            phone_number = generate_phone_number()
            parent_id = generate_parent_id()
            level = generate_level()
            group = generate_group()

            student = {
                'CNE': cne,
                'First_Name': first_name,
                'Last_Name': last_name,
                'Gender': gender,
                'Email': email,
                'Phone_Number': phone_number,
                'Parent_ID': parent_id,
                'Field_ID': field_id,
                'Level': level,
                'Group': group
            }
            if len(Etudiant.objects.all()) < 50:
                E = Etudiant(None, student['CNE'], student['First_Name'], student['Last_Name'], student['Gender'],
                             student['Email'], student['Phone_Number'], student['Group'], student['Level'],
                             student['Parent_ID'], student['Field_ID'])
                E.save()
                students.append(student)


# Fonction pour générer un nom de cours aléatoire
def generate_course_name():
    course_prefixes = ['Web_dev', 'Science', 'History', 'English', 'Art', 'Physics', 'programming']
    return random.choice(course_prefixes)


# Fonction pour générer une heure aléatoire dans la journée
def generate_random_time():
    start = datetime.datetime.strptime('08:00', '%H:%M').time()
    end = datetime.datetime.strptime('18:00', '%H:%M').time()
    random_minutes = random.randint(0, int((end.hour - start.hour) * 60 + (end.minute - start.minute)))
    random_time = datetime.datetime.combine(datetime.datetime.today(), start) + datetime.timedelta(
        minutes=random_minutes)
    return random_time.time()


def courses_list():
    # Générer les informations pour les cours
    courses = []
    for _ in range(30):
        course_name = generate_course_name()
        field = random.choice(fields)
        start_time = generate_random_time()
        duration_minutes = random.randint(30, 180)  # Random duration between 30 and 180 minutes
        end_time = (datetime.datetime.combine(datetime.datetime.today(), start_time) + datetime.timedelta(
            minutes=duration_minutes)).time()
        course = {
            'Course_Name': course_name,
            'Start_Time': start_time,
            'End_Time': end_time,
            'Field': field['name']
        }
        courses.append(course)
        if (len(Cour.objects.all()) < 30):
            cour = Cour(None, course['Course_Name'], course['Start_Time'], course['End_Time'])
            cour.save()


def generate_random_date_in_2023():
    start_date = datetime.date(2023, 1, 1)
    end_date = datetime.date(2023, 12, 31)
    random_days = random.randint(0, (end_date - start_date).days)
    random_date = start_date + datetime.timedelta(days=random_days)
    return random_date


def program_course():
    listeSalle = Salle.objects.all()
    i = 0
    if len(Salle_Cour.objects.all()) < 30:

        for c in Cour.objects.all():
            c = Salle_Cour(None, c.id, listeSalle[i].id, generate_random_date_in_2023())
            i = i + 1
            c.save()
        if i == 20:
            i = 0


def generateFieldsList():
    for i in range(5):
        if len(Filiere.objects.all()) < 4:
            f = Filiere(None, fields[i]['name'])
            f.save()


def generateClasses():
    classes = []
    # Define the flats
    flats = [1, 2, 3, 4, 5]
    # Generate classes for each flat
    for flat in flats:
        # Add TP class (CC) for each flat
        tp_class_id = "CC" + str(flat)
        tp_class_data = {
            "id": tp_class_id,
            "flat": flat,
            "type": "TP",
            "camera_reference": tp_class_id
        }
        classes.append(tp_class_data)
        # Add regular classes (A, B, C) for each flat
        for i in range(3):
            class_id = chr(65 + i) + str(flat)
            class_data = {
                "id": class_id,
                "flat": flat,
                "type": "COUR",
                "camera_reference": class_id
            }
            classes.append(class_data)
    if len(Salle.objects.all()) >= 20:
        pass
    else:
        for c in classes:
            s = Salle(None, c['id'], c['flat'], c['type'])
            s.save()


def DisplayStudentsList(request):
    pass


def findEncodings(images):
    encodingsList = []
    for img in images:
        img = cv2.cvtColor(img, cv2.COLOR_BGR2RGB)
        encode = face_recognition.face_encodings(img)[0]
        encodingsList.append(encode)
    return encodingsList


def findByCne(cne):
    students = Etudiant.objects.all()
    for e in students:
        if (e.cne == cne):
            return e
    return None


def markAttendance(cne):
    with open('attendance.csv', 'w+') as f:
        myDataList = f.readlines()
        nameList = []
        for line in myDataList:
            entry = line.split(',')
            nameList.append(entry[0])
        if cne not in nameList:
            print(cne + 'not in the list')
            now = timezone.now()
            item = findByCne(cne)
            list = Cour.objects.all()
            for l in list:
                if is_time_in_range(now.strftime('%H:%M'), l.heureDebut.strftime('%H:%M'),
                                    l.heureFin.strftime('%H:%M')):
                    coursename = l.nomCour
                    break
            print(coursename)
            dtString = now.strftime('%H:%M')
            data=[]
            data.append(item.cne)
            data.append(item.nom)
            data.append(item.prenom)
            data.append(item.groupe)
            data.append(item.filiere.nomeF)
            data.append(item.niveau)
            data.append(dtString)
            data.append(coursename)
            with open('output.csv', 'a+', newline='') as file:
                writer = csv.writer(file)
                writer.writerow(data)
        f.writelines(f'\n{cne},{item.nom},{item.prenom},{item.groupe},{item.filiere.nomeF},{item.niveau},{dtString},{coursename}')
        print('ecriture effectuee')


def etudiant_list(request):
    etudiants = Etudiant.objects.all()
    paginator = Paginator(etudiants, 12)
    p = request.GET.get("page",1)
    try:
        etudiantsPage = paginator.page(p)
    except PageNotAnInteger:
        etudiantsPage = paginator.page(1)
    except EmptyPage:
        etudiantsPage = paginator.page(1)
    return render(request, 'etudiant_list.html', {'etudiants':etudiantsPage, 'page':p})

def prof_list(request):
    professors = Professeur.objects.all()
    paginator = Paginator(professors, 12)
    p = request.GET.get("page",1)
    try:
        professorsPage = paginator.page(p)
    except PageNotAnInteger:
        professorsPage = paginator.page(1)
    except EmptyPage:
        professorsPage = paginator.page(1)
    return render(request, 'prof_list.html', {'professors':professorsPage, 'page':p})



def searchProf(request):
    query= request.GET.get('keyword')
    Professors = Professeur.objects.filter(nom__contains = query)
    paginator = Paginator(Professors, 12)
    p = request.GET.get("page",1)
    try:
        professorsPage = paginator.page(p)
    except PageNotAnInteger:
        professorsPage = paginator.page(1)
    except EmptyPage:
        professorsPage = paginator.page(1)
    return render(request, 'prof_list.html', {'professors':professorsPage, 'page':p})
    

def searchEtud(request):
    query= request.GET.get('keyword')
    etudiants = Etudiant.objects.filter(nom__contains = query)
    paginator = Paginator(etudiants, 12)
    p = request.GET.get("page",1)
    try:
        etudiantsPage = paginator.page(p)
    except PageNotAnInteger:
        etudiantsPage = paginator.page(1)
    except EmptyPage:
        etudiantsPage = paginator.page(1)
    return render(request, 'prof_list.html', {'etudiants':etudiantsPage, 'page':p})

def excel_view(request):
    p = request.GET.get("page",1)
    groupe = 'GR'+str(p)
    file_path = './fiche.xlsx'
    try:
        df = pd.read_excel(file_path)
        df.columns = ['CNE', 'PRENOM', 'NOM','Genre', 'Filiere', 'Horaire','Presence', 'Matiere']
        html_table = df.to_html(classes='table table-striped')
        return render(request, 'data.html', {'data': html_table,'groupe':p})
    except FileNotFoundError:
        return render(request, 'data.html', {'data': 'File not found'})
    
    
def chart_view(request):
    # groupe = 'GR'+str(grp)
    # df = pd.read_excel('./'+str(groupe)+'.xlsx')
    df = pd.read_excel('./fiche.xlsx')
    fig, ax = plt.subplots(figsize=(12, 6))

    df = pd.read_excel('./fiche.xlsx')
    plt.figure(figsize=(10, 6))

    distinct_values = df.iloc[:,4].unique()
    filtered_rows = df[df.iloc[:, 6] == 'Absent(e)']
    distinct_values = filtered_rows.iloc[:, 4].value_counts()
    
    distinct_names = distinct_values.index
    distinct_counts = distinct_values.values

    plt.bar(distinct_names, distinct_counts)
    plt.xlabel('filiere')
    plt.ylabel('nombre d\'absences')
    plt.title('Attendance')
    plt.savefig('App/static/App/img/Absence.png')

    df2 = pd.read_excel('./fiche.xlsx')
    plt.figure(figsize=(10, 6))

    distinct_values = df2.iloc[:,4].unique()
    filtered_rows = df2[df2.iloc[:, 6] == 'Absent(e)']
    distinct_values = filtered_rows.iloc[:, 3].value_counts()
    distinct_names = distinct_values.index
    distinct_counts = distinct_values.values

    plt.pie(distinct_counts, labels=distinct_names, autopct='%1.1f%%')
    plt.title('Genre Distribution')
    plt.savefig('App/static/App/img/AbsenceP.png')


    df3 = pd.read_excel('./fiche.xlsx')
    plt.figure(figsize=(10, 6))

    distinct_values = df3.iloc[:,4].unique()
    filtered_rows = df3[df3.iloc[:, 6] == 'Absent(e)']
    distinct_values = filtered_rows.iloc[:, 7].value_counts()
    
    distinct_names = distinct_values.index
    distinct_counts = distinct_values.values

    plt.bar(distinct_names, distinct_counts)
    plt.xlabel('Matiere')
    plt.ylabel('nombre d\'absences')
    plt.title('presence par matiere')
    plt.savefig('App/static/App/img/AbsenceG.png')

    return render(request, 'chart.html', {'chart': 'App/img/Absence.png','chartP': 'App/img/AbsenceP.png','chartG': 'App/img/AbsenceG.png'})

# Create your views here.
def homepage(request):
    return render(request, 'index.html')


def attendance_page(request):
    return render(request, 'absence.html')


def attendance(request):
    facerecognition()
    generateExcelFile()
    return render(request, 'index.html')

def generate_cne():
    return ''.join(random.choices(string.digits, k=10))



def findByCne(cne):
    students = Etudiant.objects.all()
    for e in students:
        if (e.cne == cne):
            return e
    return None

def groupStudents(grp):
    l=Etudiant.objects.all()
    output=[]
    if(l.groupe==grp):
        output.append(l)
    return output

def markAttendance(name,matiere):
    import MySQLdb
    ListPresence = []
    ListEtudiant = []
    db = MySQLdb.connect(host="localhost",
                        user="root",
                        passwd="",
                        db="etablissement")
    with open('attendance.csv','w+') as f:
        myDataList = f.readlines()
        nameList = []
        for line in myDataList:
            entry = line.split(',')
            nameList.append(entry[0])
        if name not in nameList:
            now = timezone.now()
            dtString = now.strftime('%H:%M:%S')
            # create a cursor object
            cursor = db.cursor()
            # execute SQL query
            cursor.execute(
                "SELECT app_cour.*,app_semaine.jour FROM app_cour, app_semaine WHERE app_cour.jour = app_semaine.id ")
            # fetch data
            data = cursor.fetchall()
            current_day = timezone.now().strftime('%A')
            for row in data:
                nameList = []
                dateDebut = row[2]
                dateFin = row[3]
                time_delta = pd.Timedelta(dtString)
                time = str(time_delta)
                hours, minutes, seconds = map(int, time.split(' ')[-1].split(':'))
                current_time = datetime.timedelta(hours=hours, minutes=minutes, seconds=seconds)
                if (current_day == row[5]):
                    print("test1")
                    if(is_between_dates(current_time,dateDebut,dateFin)):
                        matiere=row[1]
                        print(matiere)
                        print("test2")
            cursor = db.cursor()
            cursor.execute(
                "SELECT cne FROM app_etudiant")
            names = cursor.fetchall()
            count=0
            for nom in names :
                if (name==nom[0].upper()):
                    count+=1
                    if name not in ListPresence :
                        ListPresence.append(name)
                else:
                    pass

            for item in ListPresence :
                print('item is'+item)
                s=findByCne(item)
                f.writelines(f'\n{item},{s.nom.upper()},{s.prenom},{s.genre},{s.filiere.nomeF},{dtString},{"Present(e)"},{matiere}')
            for nom in names :
                if nom[0].upper() not in ListPresence :
                    s = findByCne(nom[0])
                    f.writelines(f'\n{nom[0].upper()},{s.nom.upper()},{s.prenom},{s.genre},{s.filiere.nomeF},{"-"},{"Absent(e)"},{matiere}')
                    print(ListPresence)

def is_between_dates(current_date, start_date, end_date):
    print(start_date, current_date, end_date)
    if start_date <= current_date <= end_date:
        return True
    else:
        return False
    
def facerecognition():
    path = 'ImagesAttendance'
    images = []
    classNames = []
    myList = os.listdir(path)

    for cls in myList:
        curImg = cv2.imread(f'{path}/{cls}')
        images.append(curImg)
        classNames.append(os.path.splitext(cls)[0])
    encodeListKnown = findEncodings(images)
    cap = cv2.VideoCapture(0)
    while True:
        success, img = cap.read()
        imgS = cv2.resize(img, (0, 0), None, 0.25, 0.25)
        img = cv2.cvtColor(imgS, cv2.COLOR_BGR2RGB)
        facesCurFrame = face_recognition.face_locations(imgS)
        encodesCurFrame = face_recognition.face_encodings(imgS, facesCurFrame)
        for encodeFace, faceLoc in zip(encodesCurFrame, facesCurFrame):
            matches = face_recognition.compare_faces(encodeListKnown, encodeFace)
            faceDis = face_recognition.face_distance(encodeListKnown, encodeFace)
            matchIndex = np.argmin(faceDis)
            if matches[matchIndex]:
                name = classNames[matchIndex].upper()
                matiere='PFA'
                y1, x2, y2, x1 = faceLoc
                cv2.rectangle(img, (x1, y1), (x2, y2), (0, 255, 0), 2)
                cv2.rectangle(img, (x1, y2 - 35), (x2, y2), (0, 255, 0), cv2.FILLED)
                cv2.putText(img, name, (x1 + 6, y2 - 6), cv2.FONT_HERSHEY_COMPLEX, 1, (255, 255, 255), 2)
                markAttendance(name,matiere)
                generateExcelFile()
                break
        cv2.imshow('webcam', img)
        cv2.waitKey(1)

def generateExcelFile():
    workbook = Workbook()

    worksheet = workbook.active
    contenu = []
    with open('attendance.csv', 'r') as f:
        lines = f.readlines()
    x = []
    for l in lines:
        x = l.split(',')
        contenu.append(x)
    for row_index, row in enumerate(contenu, start=1):
        for col_index, value in enumerate(row, start=1):
            cell = worksheet.cell(row=row_index, column=col_index)
            cell.value = value
    workbook.save('fiche.xlsx')

def generate_pdf(request):
    image_paths = ['App/img/Absence.png', 'App/img/AbsenceP.png', 'App/img/AbsenceG.png']
    excel_file_path = './fiche.xlsx'

    pdf_file = BytesIO()
    doc = SimpleDocTemplate(pdf_file, pagesize=letter)

    elements = []

    header_text = "EDUCAT Le "+ str(timezone.now().strftime("%Y-%m-%d"))
    header_style = "Helvetica-Bold"
    header_color = colors.HexColor("#06BBCC")
    header = Paragraph(header_text, style=ParagraphStyle(name='Header', fontName=header_style, textColor=header_color, fontSize=16,  leading=24))
    elements.append(header)

    images = []
    for image_path in image_paths:
        image_absolute_path = finders.find(image_path)
        image = Image(image_absolute_path, width=500, height=300) 
        images.append(image)
    wb = load_workbook(excel_file_path)
    ws = wb.active
    excel_data = ws.values
    table_data = []
    for row in excel_data:
        table_data.append(row)
    table = Table(table_data)

    # Apply table style
    table.setStyle(TableStyle([
        ('BACKGROUND', (0, 0), (-1, 0), colors.HexColor('#06BBCC')),
        ('TEXTCOLOR', (0, 0), (-1, 0), colors.white),
        ('ALIGN', (0, 0), (-1, -1), 'CENTER'),
        ('FONTNAME', (0, 0), (-1, 0), 'Helvetica-Bold'),
        ('FONTSIZE', (0, 0), (-1, 0), 12),
        ('BOTTOMPADDING', (0, 0), (-1, 0), 12),
        ('BACKGROUND', (0, 1), (-1, -1), colors.white), 
        ('BOX', (0, 0), (-1, -1), 1, colors.black),
        ('GRID', (0, 0), (-1, -1), 1, colors.black),
    ]))

    # Add the images and table to the PDF elements list
    elements.append(table)
    elements.extend(images)
    

    # Build the PDF document
    doc.build(elements)

    # Return the PDF file as a response
    response = HttpResponse(content_type='application/pdf')
    response['Content-Disposition'] = 'attachment; filename="generated_pdf.pdf"'
    response.write(pdf_file.getvalue())

    return response



@csrf_exempt
def send_email_Prof(request):
    if request.method == 'POST':
        email_form = EmailForm(request.POST)
        if email_form.is_valid():
            student_id = request.POST.get('student_id')
            email_body = email_form.cleaned_data['email_body']

        # Retrieve student details based on the ID
        student = Etudiant.objects.get(id=student_id)

        # Send the email
        send_mail(
            'EMAIL',
            email_body,
            'educatetablissement@gmail.com',
            # ['youneshssine@gmail.com'],
            [student.email],
            fail_silently=False,
        )
    else:
        email_form = EmailForm()
    return redirect('etudiant_list')


@csrf_exempt
def send_email_Etud(request):
    if request.method == 'POST':
        email_form = EmailForm(request.POST)
        if email_form.is_valid():
            student_id = request.POST.get('student_id')
            email_body = email_form.cleaned_data['email_body']

        # Retrieve student details based on the ID
        student = Etudiant.objects.get(id=student_id)
        # Send the email
        send_mail(
            'EMAIL',
            email_body,
            'educatetablissement@gmail.com',
            # ['youneshssine@gmail.com'],
            [student.email],
            fail_silently=False,
        )
    else:
        email_form = EmailForm()
    return redirect('etudiant_list')

@csrf_exempt
def send_email_Prof(request):
    if request.method == 'POST':
        email_form = EmailForm(request.POST)
        if email_form.is_valid():
            professeur_id = request.POST.get('professeur_id')
            email_body = email_form.cleaned_data['email_body']

        professor = Professeur.objects.get(id=professeur_id)
        send_mail(
            'EMAIL',
            email_body,
            'educatetablissement@gmail.com',
            # ['youneshssine@gmail.com'],
            [professor.email],
            fail_silently=False,
        )
    else:
        email_form = EmailForm()

    return redirect('prof_list')


